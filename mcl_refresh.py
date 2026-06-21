"""
Mwananchi Credit Ltd — Portfolio Excel Refresh
Connects to SQL Server, runs all stored procedures, writes formatted Excel.
Run via MCL_Excel_Refresh.command or directly:
    python3 mcl_refresh.py [YYYY-MM-DD]
"""

import os
import sys
import pymssql
from datetime import datetime, date
from openpyxl import Workbook
from openpyxl.styles import (
    PatternFill, Font, Alignment, Border, Side, GradientFill
)
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import ColorScaleRule, DataBarRule
from openpyxl.chart import BarChart, Reference, PieChart
from openpyxl.chart.series import DataPoint

# ── Credentials ───────────────────────────────────────────────
SQL_SERVER   = os.environ.get("SQL_SERVER",   "173.249.59.109")
SQL_PORT     = int(os.environ.get("SQL_PORT", "1433"))
SQL_DATABASE = os.environ.get("SQL_DATABASE", "mwananchidy365")
SQL_USER     = os.environ.get("SQL_USER",     "monday")
SQL_PASSWORD = os.environ.get("SQL_PASSWORD", "")

# ── Brand colours ─────────────────────────────────────────────
NAVY   = "020F71"
ORANGE = "F7941D"
WHITE  = "FFFFFF"
LIGHT  = "F5F4F0"
GREEN  = "276749"
RED    = "C53030"
AMBER  = "C9A84C"
GREY   = "6B6560"
BORDER_CLR = "E2DDD6"

def navy_fill():  return PatternFill("solid", fgColor=NAVY)
def orange_fill(): return PatternFill("solid", fgColor=ORANGE)
def light_fill():  return PatternFill("solid", fgColor="EEF3F9")
def alt_fill():    return PatternFill("solid", fgColor="F9F8F5")
def green_fill():  return PatternFill("solid", fgColor="C6F6D5")
def red_fill():    return PatternFill("solid", fgColor="FED7D7")
def amber_fill():  return PatternFill("solid", fgColor="FEFCBF")

def header_font(size=11): return Font(name="Calibri", bold=True, color=WHITE, size=size)
def body_font(size=10):   return Font(name="Calibri", size=size)
def title_font(size=14):  return Font(name="Calibri", bold=True, color=NAVY, size=size)
def sub_font(size=10):    return Font(name="Calibri", color=GREY, size=size)
def num_font(size=10):    return Font(name="Calibri", size=size)

def thin_border():
    s = Side(style="thin", color=BORDER_CLR)
    return Border(left=s, right=s, top=s, bottom=s)

def fmt_kes(n):
    if n is None: return "—"
    try:
        v = float(n)
        return f"{v:,.2f}"
    except:
        return str(n)

def fmt_pct(n):
    if n is None: return "—"
    try: return f"{float(n):.2f}%"
    except: return str(n)

def par_fill(pct):
    try:
        p = float(pct)
        if p <= 5:   return green_fill()
        if p <= 20:  return PatternFill("solid", fgColor="C6F6D5")
        if p <= 50:  return amber_fill()
        if p <= 100: return red_fill()
        return PatternFill("solid", fgColor="C53030")
    except:
        return None

# ── DB helpers ────────────────────────────────────────────────
def get_conn():
    return pymssql.connect(
        server=SQL_SERVER, port=SQL_PORT,
        user=SQL_USER, password=SQL_PASSWORD,
        database=SQL_DATABASE, login_timeout=20, as_dict=True
    )

def run_sp(sp_name, as_at_date=None):
    conn = get_conn()
    cur  = conn.cursor()
    if as_at_date:
        try:
            cur.execute(f"EXEC {sp_name} @AsAtDate = %s", (as_at_date,))
            rows = cur.fetchall()
        except Exception:
            conn.close()
            conn = get_conn()
            cur  = conn.cursor()
            cur.execute(f"EXEC {sp_name}")
            rows = cur.fetchall()
    else:
        cur.execute(f"EXEC {sp_name}")
        rows = cur.fetchall()
    conn.close()
    return rows

def run_scalar(sp_name):
    conn = get_conn()
    cur  = conn.cursor()
    cur.execute(f"EXEC {sp_name}")
    row = cur.fetchone()
    conn.close()
    return row

# ── Sheet helpers ─────────────────────────────────────────────
def set_col_width(ws, col, width):
    ws.column_dimensions[get_column_letter(col)].width = width

def write_sheet_title(ws, row, title, subtitle, col_span):
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=col_span)
    c = ws.cell(row=row, column=1, value=title)
    c.font = Font(name="Calibri", bold=True, color=NAVY, size=14)
    c.alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[row].height = 22
    row += 1
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=col_span)
    c = ws.cell(row=row, column=1, value=subtitle)
    c.font = sub_font(9)
    c.alignment = Alignment(horizontal="left")
    ws.row_dimensions[row].height = 14
    return row + 2  # next data row

def write_header_row(ws, row, headers, col_start=1):
    for i, h in enumerate(headers):
        c = ws.cell(row=row, column=col_start+i, value=h)
        c.fill  = navy_fill()
        c.font  = header_font(10)
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border = thin_border()
    ws.row_dimensions[row].height = 18
    return row + 1

def write_data_row(ws, row, values, is_alt=False):
    fill = alt_fill() if is_alt else PatternFill("solid", fgColor=WHITE)
    for i, v in enumerate(values):
        c = ws.cell(row=row, column=1+i, value=v)
        c.font   = body_font(10)
        c.fill   = fill
        c.border = thin_border()
        if isinstance(v, (int, float)) and not isinstance(v, bool):
            c.alignment = Alignment(horizontal="right")
        else:
            c.alignment = Alignment(horizontal="left")
    return row + 1

def write_total_row(ws, row, values):
    for i, v in enumerate(values):
        c = ws.cell(row=row, column=1+i, value=v)
        c.fill  = PatternFill("solid", fgColor="020F71")
        c.font  = Font(name="Calibri", bold=True, color=WHITE, size=10)
        c.border = thin_border()
        if isinstance(v, (int, float)) and not isinstance(v, bool):
            c.alignment = Alignment(horizontal="right")
    ws.row_dimensions[row].height = 16


# ── Sheet 1: Summary ──────────────────────────────────────────
def build_summary_sheet(wb, as_at_date, branch="ALL"):
    ws = wb.active
    ws.title = "Summary"

    # Logo area
    ws.merge_cells("A1:F1")
    c = ws["A1"]
    c.value = "MWANANCHI CREDIT LTD — PORTFOLIO DASHBOARD"
    c.fill  = navy_fill()
    c.font  = Font(name="Calibri", bold=True, color=WHITE, size=16)
    c.alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[1].height = 36

    ws.merge_cells("A2:F2")
    c = ws["A2"]
    branch_tag = f"    |    Branch: {branch}" if branch and branch.upper() != "ALL" else "    |    All Branches"
    c.value = f"As at Date: {as_at_date}{branch_tag}    |    Database: {SQL_DATABASE}    |    Generated: {datetime.now().strftime('%d %b %Y %H:%M')}"
    c.fill  = PatternFill("solid", fgColor=ORANGE)
    c.font  = Font(name="Calibri", bold=True, color=WHITE, size=10)
    c.alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[2].height = 18

    # KPI section
    kpi_row = 4
    ws.merge_cells(f"A{kpi_row}:F{kpi_row}")
    c = ws.cell(row=kpi_row, column=1, value="KEY PERFORMANCE INDICATORS")
    c.font = Font(name="Calibri", bold=True, color=NAVY, size=12)
    ws.row_dimensions[kpi_row].height = 20

    # Fetch KPI data
    kpi_labels = []
    kpi_values = []
    kpi_subs   = []
    kpi_colors = []

    try:
        r = run_scalar("sp_ActiveLoanCount")
        kpi_labels.append("Active Loans");       kpi_values.append(r["ActiveLoanCount"] if r else "—"); kpi_subs.append("Total active accounts"); kpi_colors.append("0A1A8A")
    except Exception as e:
        kpi_labels.append("Active Loans");       kpi_values.append(f"Error: {e}"); kpi_subs.append(""); kpi_colors.append(RED)

    try:
        r = run_scalar("sp_TotalLoanBalance")
        v = r["PrincipalOutstandingBalance"] if r else 0
        kpi_labels.append("Total Loan Balance"); kpi_values.append(v); kpi_subs.append("KES"); kpi_colors.append(ORANGE)
    except Exception as e:
        kpi_labels.append("Total Loan Balance"); kpi_values.append(f"Error: {e}"); kpi_subs.append(""); kpi_colors.append(RED)

    try:
        r = run_scalar("sp_LoanPrincipalBalance")
        v = r["PrincipalOutstandingBalance"] if r else 0
        kpi_labels.append("Principal Outstanding"); kpi_values.append(v); kpi_subs.append("KES"); kpi_colors.append(GREEN)
    except Exception as e:
        kpi_labels.append("Principal Outstanding"); kpi_values.append(f"Error: {e}"); kpi_subs.append(""); kpi_colors.append(RED)

    try:
        r = run_scalar("sp_Portfoliainarrears")
        kpi_labels.append("Amount in Arrears");  kpi_values.append(r["TotalAmountInArrears"] if r else 0); kpi_subs.append("KES"); kpi_colors.append(RED)
        kpi_labels.append("Loans in Arrears");   kpi_values.append(r["LoanCount"] if r else 0);           kpi_subs.append("accounts overdue"); kpi_colors.append("C05621")
        kpi_labels.append("Portfolio at Risk");  kpi_values.append(r["PercentageInArrears"] if r else 0); kpi_subs.append("PAR %"); kpi_colors.append("553C9A")
    except Exception as e:
        for lbl in ["Amount in Arrears","Loans in Arrears","Portfolio at Risk"]:
            kpi_labels.append(lbl); kpi_values.append(f"Error: {e}"); kpi_subs.append(""); kpi_colors.append(RED)

    # Write KPI cards (2 rows of 3)
    card_row = kpi_row + 1
    for i, (lbl, val, sub, col) in enumerate(zip(kpi_labels, kpi_values, kpi_subs, kpi_colors)):
        cc = (i % 3) + 1           # column 1-3
        rr = card_row + (i // 3) * 3  # every 3 items = new row group
        ws.merge_cells(start_row=rr, start_column=cc*2-1, end_row=rr, end_column=cc*2)
        c = ws.cell(row=rr, column=cc*2-1, value=lbl.upper())
        c.fill = PatternFill("solid", fgColor=col)
        c.font = Font(name="Calibri", bold=True, color=WHITE, size=9)
        c.alignment = Alignment(horizontal="left", vertical="center")
        ws.row_dimensions[rr].height = 14

        ws.merge_cells(start_row=rr+1, start_column=cc*2-1, end_row=rr+1, end_column=cc*2)
        c = ws.cell(row=rr+1, column=cc*2-1, value=val)
        c.fill = PatternFill("solid", fgColor="F9F8F5")
        c.font = Font(name="Calibri", bold=True, color=NAVY, size=14)
        c.alignment = Alignment(horizontal="left", vertical="center")
        if isinstance(val, (int, float)):
            c.number_format = '#,##0.00' if sub == "PAR %" or sub == "KES" else '#,##0'
        ws.row_dimensions[rr+1].height = 22

        ws.merge_cells(start_row=rr+2, start_column=cc*2-1, end_row=rr+2, end_column=cc*2)
        c = ws.cell(row=rr+2, column=cc*2-1, value=sub)
        c.fill = PatternFill("solid", fgColor="EEF3F9")
        c.font = sub_font(8)
        c.alignment = Alignment(horizontal="left")
        ws.row_dimensions[rr+2].height = 12

    # Navigation section
    nav_row = card_row + 8
    ws.merge_cells(f"A{nav_row}:F{nav_row}")
    c = ws.cell(row=nav_row, column=1, value="SHEET GUIDE")
    c.font = Font(name="Calibri", bold=True, color=NAVY, size=11)
    ws.row_dimensions[nav_row].height = 18

    sheets_info = [
        ("Branch Arrears",    "sp_PortfolioInArrearsByBranch",   "Portfolio in arrears by branch — loan count, arrears amount, balance, PAR%"),
        ("Loan Products",     "sp_PortfolioByLoanProduct",        "Portfolio breakdown by loan product type"),
        ("Loan Category",     "sp_PortfolioByLoanCategory",       "Portfolio by credit classification (Performing, Watch, Substandard, Doubtful, Loss)"),
        ("Active Loans",      "sp_ActiveLoanCount",               "Total count of active loans"),
        ("Total Balance",     "sp_TotalLoanBalance",              "Total loan balance outstanding"),
        ("Principal Balance", "sp_LoanPrincipalBalance",          "Principal outstanding balance"),
        ("Portfolio Arrears", "sp_Portfoliainarrears",            "Overall portfolio in arrears — total amount, count, PAR%"),
    ]
    nav_row += 1
    write_header_row(ws, nav_row, ["Sheet Name","Stored Procedure","Description"])
    nav_row += 1
    for i, (sh, sp, desc) in enumerate(sheets_info):
        write_data_row(ws, nav_row, [sh, sp, desc], is_alt=(i%2==1))
        nav_row += 1

    # Column widths
    for col, w in [(1,22),(2,22),(3,22),(4,22),(5,22),(6,22)]:
        set_col_width(ws, col, w)

    # Freeze panes
    ws.freeze_panes = "A3"


# ── Generic raw-data sheet builder ────────────────────────────
def build_sp_sheet(wb, sheet_name, sp_name, as_at_date, title, subtitle,
                   col_formats=None, par_col=None, filter_total_col=None,
                   branch_filter=None):
    """
    col_formats: dict {col_index: 'kes'|'pct'|'int'|'text'}
    par_col: 1-based column index to apply PAR colour fill
    filter_total_col: column key name whose value 'TOTAL' rows should be removed
    """
    ws = wb.create_sheet(sheet_name)
    ws.sheet_view.showGridLines = False

    # Title banner
    next_row = 1
    ws.merge_cells(f"A{next_row}:Z{next_row}")
    c = ws.cell(row=next_row, column=1, value=f"MWANANCHI CREDIT LTD — {title.upper()}")
    c.fill = navy_fill()
    c.font = Font(name="Calibri", bold=True, color=WHITE, size=13)
    c.alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[next_row].height = 28
    next_row += 1

    ws.merge_cells(f"A{next_row}:Z{next_row}")
    c = ws.cell(row=next_row, column=1, value=f"{subtitle}  |  SP: {sp_name}  |  As at: {as_at_date}  |  Generated: {datetime.now().strftime('%d %b %Y %H:%M')}")
    c.fill = PatternFill("solid", fgColor=ORANGE)
    c.font = Font(name="Calibri", color=WHITE, size=9)
    c.alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[next_row].height = 14
    next_row += 2

    # Fetch data
    try:
        rows = run_sp(sp_name, as_at_date=as_at_date)
    except Exception as e:
        ws.cell(row=next_row, column=1, value=f"ERROR connecting to {sp_name}: {e}")
        ws.cell(row=next_row, column=1).font = Font(color=RED, bold=True, size=11)
        return ws

    if not rows:
        ws.cell(row=next_row, column=1, value=f"No data returned from {sp_name} for date {as_at_date}")
        return ws

    # Filter TOTAL rows if requested
    if filter_total_col:
        rows = [r for r in rows if str(r.get(filter_total_col, "")).upper() != "TOTAL"]

    # Apply branch filter (post-fetch) — only if column exists in result
    if branch_filter:
        branch_val, branch_col = branch_filter
        if branch_val and branch_val.upper() != "ALL" and rows and branch_col in rows[0]:
            rows = [r for r in rows if str(r.get(branch_col, "")).strip().upper() == branch_val.upper()]
            if not rows:
                ws.cell(row=4, column=1,
                    value=f"No data found for branch '{branch_val}' in {sp_name}.")
                return ws

    # Headers
    headers = list(rows[0].keys())
    header_row = next_row
    write_header_row(ws, header_row, headers)
    next_row = header_row + 1

    # Data rows
    for i, row in enumerate(rows):
        values = []
        for h in headers:
            v = row.get(h)
            if isinstance(v, (datetime, date)):
                v = v.strftime("%Y-%m-%d")
            values.append(v)
        data_row_num = next_row

        is_alt = (i % 2 == 1)
        fill = alt_fill() if is_alt else PatternFill("solid", fgColor=WHITE)

        for j, (h, v) in enumerate(zip(headers, values)):
            c = ws.cell(row=data_row_num, column=j+1, value=v)
            c.font   = body_font(10)
            c.border = thin_border()

            # PAR colour column
            if par_col and (j+1) == par_col and v is not None:
                pf = par_fill(v)
                c.fill = pf if pf else fill
                c.font = Font(name="Calibri", bold=True, size=10)
                try:
                    c.number_format = '0.00"%"'
                except:
                    pass
            else:
                c.fill = fill

            # Number formatting
            fmt = (col_formats or {}).get(j+1, "")
            if fmt == "kes" and v is not None:
                try:
                    c.value = float(v)
                    c.number_format = '#,##0.00'
                    c.alignment = Alignment(horizontal="right")
                except:
                    pass
            elif fmt == "pct" and v is not None:
                try:
                    c.value = float(v)
                    c.number_format = '0.00'
                    c.alignment = Alignment(horizontal="right")
                except:
                    pass
            elif fmt == "int" and v is not None:
                try:
                    c.value = int(v)
                    c.number_format = '#,##0'
                    c.alignment = Alignment(horizontal="right")
                except:
                    pass
            else:
                c.alignment = Alignment(horizontal="left" if isinstance(v, str) else "right")

        ws.row_dimensions[data_row_num].height = 15
        next_row += 1

    # Totals row (sum numeric columns)
    total_values = ["TOTAL"] + [""] * (len(headers)-1)
    for j, h in enumerate(headers[1:], 1):
        try:
            total_values[j] = sum(
                float(r.get(h, 0) or 0)
                for r in rows
                if r.get(h) is not None
            )
        except:
            pass
    write_total_row(ws, next_row, total_values)
    next_row += 2

    # Auto-width columns
    for j, h in enumerate(headers):
        max_len = max(len(str(h)), max((len(str(r.get(h, "") or "")) for r in rows), default=0))
        ws.column_dimensions[get_column_letter(j+1)].width = min(max_len + 4, 30)

    ws.freeze_panes = f"A{header_row+1}"
    return ws


# ── Loan Category sheet (Performing separated from arrears) ───
def build_category_sheet(wb, as_at_date, branch="ALL"):
    """
    Loan Category sheet with two sections:
      - ARREARS CATEGORIES  (Watch 1, Watch 2, Substandard, Doubtful, Loss)
      - PERFORMING — CURRENT (not in arrears, shown separately at bottom)
    """
    ws = wb.create_sheet("Loan Category")
    ws.sheet_view.showGridLines = False

    # Banner
    ws.merge_cells("A1:Z1")
    c = ws.cell(row=1, column=1, value="MWANANCHI CREDIT LTD — PORTFOLIO BY LOAN CATEGORY")
    c.fill = navy_fill(); c.font = Font(name="Calibri", bold=True, color=WHITE, size=13)
    c.alignment = Alignment(horizontal="left", vertical="center"); ws.row_dimensions[1].height = 28

    ws.merge_cells("A2:Z2")
    c = ws.cell(row=2, column=1,
        value=f"Credit classification breakdown  |  sp_PortfolioByLoanCategory  |  As at: {as_at_date}  |  {datetime.now().strftime('%d %b %Y %H:%M')}")
    c.fill = orange_fill(); c.font = Font(name="Calibri", color=WHITE, size=9)
    c.alignment = Alignment(horizontal="left", vertical="center"); ws.row_dimensions[2].height = 14

    # Fetch
    try:
        all_rows = run_sp("sp_PortfolioByLoanCategory", as_at_date=as_at_date)
    except Exception as e:
        ws.cell(row=4, column=1, value=f"ERROR: {e}").font = Font(color=RED, bold=True, size=11)
        return ws

    if not all_rows:
        ws.cell(row=4, column=1, value="No data returned for this date.")
        return ws

    # Remove TOTAL rows
    all_rows = [r for r in all_rows if str(r.get("LoanCategory","")).upper() != "TOTAL"]

    # Apply branch filter if column exists
    if branch and branch.upper() != "ALL" and all_rows and "BranchName" in all_rows[0]:
        all_rows = [r for r in all_rows if str(r.get("BranchName","")).strip().upper() == branch.upper()]
    performing = [r for r in all_rows if str(r.get("LoanCategory","")).strip().upper() == "PERFORMING"]
    arrears_rows = [r for r in all_rows if str(r.get("LoanCategory","")).strip().upper() != "PERFORMING"]

    headers = list(all_rows[0].keys())
    col_formats = {}
    par_col = None
    for j, h in enumerate(headers):
        hl = h.lower()
        if j == 0: col_formats[j+1] = "text"
        elif "count" in hl or "loans" in hl: col_formats[j+1] = "int"
        elif "percentage" in hl or "par" in hl or "rate" in hl:
            col_formats[j+1] = "pct"; par_col = j+1
        elif "amount" in hl or "balance" in hl or "arrears" in hl:
            col_formats[j+1] = "kes"

    def write_section(ws, start_row, rows, section_label, label_color):
        # Section label
        ws.merge_cells(start_row=start_row, start_column=1, end_row=start_row, end_column=len(headers))
        c = ws.cell(row=start_row, column=1, value=section_label)
        c.fill = PatternFill("solid", fgColor=label_color)
        c.font = Font(name="Calibri", bold=True, color=WHITE, size=10)
        c.alignment = Alignment(horizontal="left", vertical="center")
        ws.row_dimensions[start_row].height = 16
        start_row += 1

        # Header row
        write_header_row(ws, start_row, headers)
        start_row += 1

        # Data rows
        for i, row in enumerate(rows):
            values = []
            for h in headers:
                v = row.get(h)
                if isinstance(v, (datetime, date)): v = v.strftime("%Y-%m-%d")
                values.append(v)
            is_alt = (i % 2 == 1)
            fill = alt_fill() if is_alt else PatternFill("solid", fgColor=WHITE)
            for j, (h, v) in enumerate(zip(headers, values)):
                c = ws.cell(row=start_row, column=j+1, value=v)
                c.font = body_font(10); c.border = thin_border()
                fmt = col_formats.get(j+1, "")
                if par_col and (j+1) == par_col and v is not None:
                    pf = par_fill(v)
                    c.fill = pf if pf else fill
                    c.font = Font(name="Calibri", bold=True, size=10)
                    try: c.value = float(v); c.number_format = '0.00'
                    except: pass
                    c.alignment = Alignment(horizontal="right")
                else:
                    c.fill = fill
                    if fmt == "kes" and v is not None:
                        try: c.value = float(v); c.number_format = '#,##0.00'; c.alignment = Alignment(horizontal="right")
                        except: pass
                    elif fmt == "int" and v is not None:
                        try: c.value = int(v); c.number_format = '#,##0'; c.alignment = Alignment(horizontal="right")
                        except: pass
                    else:
                        c.alignment = Alignment(horizontal="left" if isinstance(v, str) else "right")
            ws.row_dimensions[start_row].height = 15
            start_row += 1

        # Totals row
        tot = ["TOTAL"] + [""] * (len(headers)-1)
        for j, h in enumerate(headers[1:], 1):
            try: tot[j] = sum(float(r.get(h, 0) or 0) for r in rows if r.get(h) is not None)
            except: pass
        write_total_row(ws, start_row, tot)
        return start_row + 2   # blank row gap

    next_row = 4
    # Section 1: Arrears categories
    next_row = write_section(ws, next_row, arrears_rows,
        "ARREARS CATEGORIES  —  Watch 1 / Watch 2 / Substandard / Doubtful / Loss", RED)

    # Section 2: Performing (current, not in arrears)
    next_row = write_section(ws, next_row, performing,
        "PERFORMING — CURRENT LOANS (NOT IN ARREARS)", GREEN)

    # Auto-width
    for j, h in enumerate(headers):
        max_len = max(len(str(h)), max((len(str(r.get(h,"") or "")) for r in all_rows), default=0))
        ws.column_dimensions[get_column_letter(j+1)].width = min(max_len + 4, 32)

    ws.freeze_panes = "A5"
    return ws


# ── Branch filter helper ──────────────────────────────────────
def apply_branch_filter(rows, branch, col="BranchName"):
    """Return only rows matching branch. Pass 'ALL' or None to keep all."""
    if not branch or branch.strip().upper() == "ALL":
        return rows
    return [r for r in rows if str(r.get(col, "")).strip().upper() == branch.strip().upper()]


# ── Main builder ──────────────────────────────────────────────
def build_workbook(as_at_date, branch="ALL"):
    wb = Workbook()
    branch_label = branch.strip().upper() if branch else "ALL"
    is_filtered  = branch_label != "ALL"

    print(f"\n  Mwananchi Credit — Excel Portfolio Refresh")
    print(f"  As at date : {as_at_date}")
    print(f"  Branch     : {branch_label}")
    print(f"  Database   : {SQL_DATABASE} @ {SQL_SERVER}")
    print()

    # Sheet 1: Summary
    print("  [1/8] Summary KPIs ...")
    build_summary_sheet(wb, as_at_date, branch=branch_label)
    print("        Done")

    # Sheet 2: Branch Arrears — filter post-fetch
    print("  [2/8] sp_PortfolioInArrearsByBranch ...")
    build_sp_sheet(
        wb,
        sheet_name="Branch Arrears",
        sp_name="sp_PortfolioInArrearsByBranch",
        as_at_date=as_at_date,
        title=f"Portfolio in Arrears by Branch{' — '+branch_label if is_filtered else ''}",
        subtitle="Branch-level arrears, loan balance and PAR percentage",
        col_formats={3:"int", 4:"kes", 5:"kes", 6:"pct"},
        par_col=6,
        filter_total_col="BranchName",
        branch_filter=(branch_label, "BranchName"),
    )
    print("        Done")

    # Sheet 3: Loan Products — filter post-fetch if SP returns BranchName
    print("  [3/8] sp_PortfolioByLoanProduct ...")
    build_sp_sheet(
        wb,
        sheet_name="Loan Products",
        sp_name="sp_PortfolioByLoanProduct",
        as_at_date=as_at_date,
        title=f"Portfolio by Loan Product{' — '+branch_label if is_filtered else ''}",
        subtitle="Arrears and balance breakdown by loan product type",
        col_formats={2:"int", 3:"kes", 4:"kes", 5:"pct"},
        par_col=5,
        filter_total_col="LoanProductType",
        branch_filter=(branch_label, "BranchName"),
    )
    print("        Done")

    # Sheet 4: Loan Category (Performing separated — not in arrears)
    print("  [4/8] sp_PortfolioByLoanCategory ...")
    build_category_sheet(wb, as_at_date, branch=branch_label)
    print("        Done")

    # Sheet 5: Active Loans (summary — no branch split)
    print("  [5/8] sp_ActiveLoanCount ...")
    build_sp_sheet(
        wb, sheet_name="Active Loans", sp_name="sp_ActiveLoanCount",
        as_at_date=None, title="Active Loan Count",
        subtitle="Total number of active loans in the portfolio",
    )
    print("        Done")

    # Sheet 6: Total Loan Balance
    print("  [6/8] sp_TotalLoanBalance ...")
    build_sp_sheet(
        wb, sheet_name="Total Balance", sp_name="sp_TotalLoanBalance",
        as_at_date=None, title="Total Loan Balance",
        subtitle="Principal outstanding balance across all active loans",
        col_formats={1:"kes"},
    )
    print("        Done")

    # Sheet 7: Principal Balance
    print("  [7/8] sp_LoanPrincipalBalance ...")
    build_sp_sheet(
        wb, sheet_name="Principal Balance", sp_name="sp_LoanPrincipalBalance",
        as_at_date=None, title="Loan Principal Balance",
        subtitle="Principal outstanding balance", col_formats={1:"kes"},
    )
    print("        Done")

    # Sheet 8: Portfolio Arrears
    print("  [8/8] sp_Portfoliainarrears ...")
    build_sp_sheet(
        wb, sheet_name="Portfolio Arrears", sp_name="sp_Portfoliainarrears",
        as_at_date=None, title="Portfolio in Arrears — Overall",
        subtitle="Total arrears, loan count in arrears, PAR percentage",
        col_formats={1:"kes", 3:"pct"},
    )
    print("        Done")

    return wb


def main():
    as_at_date   = sys.argv[1] if len(sys.argv) > 1 else datetime.now().strftime("%Y-%m-%d")
    branch       = sys.argv[2] if len(sys.argv) > 2 else "ALL"
    branch_label = branch.strip().upper()
    # Optional 3rd arg: custom output filename (used by auto-refresh daemon)
    # e.g.  python3 mcl_refresh.py 2026-06-20 ALL MCL_Portfolio_LIVE.xlsx
    output_override = sys.argv[3] if len(sys.argv) > 3 else None

    if output_override:
        filename = output_override
    elif branch_label == "ALL":
        filename = f"MCL_Portfolio_{as_at_date}.xlsx"
    else:
        filename = f"MCL_Portfolio_{as_at_date}_{branch_label}.xlsx"

    out_path = os.path.join(os.path.expanduser("~"), "mwananchi-portfolio", filename)

    wb = build_workbook(as_at_date, branch=branch_label)
    wb.save(out_path)
    print(f"\n  Excel saved to:\n  {out_path}\n")
    return out_path


if __name__ == "__main__":
    main()
